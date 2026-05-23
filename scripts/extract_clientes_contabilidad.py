"""
Extrae un resumen JSON del workbook "Datos Clientes Y Contabilidad.xlsx".

Uso:
    python scripts/extract_clientes_contabilidad.py --path "C:\\ruta\\archivo.xlsx"
"""

from __future__ import annotations

import argparse
import json
import math
import os
import re
import subprocess
import sys
import tempfile
import unicodedata
from collections import Counter
from contextlib import contextmanager
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def _copy_locked_file(src: Path, dst: Path) -> None:
    """
    Copia un archivo respetando los share modes de Windows. Excel abre el .xlsx
    con FILE_SHARE_READ, así que el `copy` nativo de Windows puede leerlo aunque
    Python `open()` falle con PermissionError.
    """
    if os.name == "nt":
        result = subprocess.run(
            ["cmd", "/c", "copy", "/Y", str(src), str(dst)],
            capture_output=True,
            text=True,
            shell=False,
        )
        if result.returncode != 0:
            raise RuntimeError(
                f"No se pudo copiar el Excel ({result.returncode}): "
                f"{(result.stderr or result.stdout).strip()}"
            )
        return

    # Fallback en sistemas no-Windows
    import shutil
    shutil.copyfile(src, dst)


@contextmanager
def open_workbook_safely(workbook_path: Path):
    """
    Carga el workbook copiándolo a una ruta temporal primero, así no fallamos
    si el usuario tiene el archivo abierto en Excel.
    """
    tmp_handle = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp_handle.close()
    tmp_path = Path(tmp_handle.name)
    try:
        _copy_locked_file(workbook_path, tmp_path)
        wb = load_workbook(tmp_path, read_only=True, data_only=True, keep_links=False)
        try:
            yield wb
        finally:
            wb.close()
    finally:
        try:
            tmp_path.unlink()
        except OSError:
            pass


DEFAULT_EXCEL_PATH = Path(
    r"C:\Users\SANTIAGO\OneDrive - uniminuto.edu\Escritorio\Datos Clientes Y Contabilidad.xlsx"
)

MONTH_NAMES = {
    1: "Ene",
    2: "Feb",
    3: "Mar",
    4: "Abr",
    5: "May",
    6: "Jun",
    7: "Jul",
    8: "Ago",
    9: "Sep",
    10: "Oct",
    11: "Nov",
    12: "Dic",
}


def to_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\xa0", " ").strip()
    return text or None


def to_number(value: Any) -> int | float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if math.isnan(value):
            return None
        return int(value) if value.is_integer() else float(value)
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    if math.isnan(parsed):
        return None
    return int(parsed) if parsed.is_integer() else parsed


def to_float(value: Any) -> float:
    parsed = to_number(value)
    return float(parsed) if parsed is not None else 0.0


def to_iso_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return None


def to_serialized(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def identifier_text(value: Any) -> str | None:
    parsed = to_number(value)
    if parsed is None:
        return to_text(value)
    if isinstance(parsed, int):
        return str(parsed)
    if float(parsed).is_integer():
        return str(int(parsed))
    return str(parsed)


def normalize_status(value: Any) -> str:
    text = to_text(value)
    return text or "Sin estado"


def month_label(dt: date | datetime | None) -> str | None:
    if dt is None:
        return None
    if isinstance(dt, datetime):
        dt = dt.date()
    return f"{MONTH_NAMES.get(dt.month, str(dt.month))} {dt.year}"


def normalize_margin_pct(value: Any) -> float:
    parsed = to_number(value)
    if parsed is None:
        return 0.0
    parsed = float(parsed)
    if abs(parsed) <= 1.5:
        return parsed * 100
    return parsed


def normalize_key(value: Any) -> str:
    text = to_text(value)
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = text.lower().replace("&", " y ")
    return re.sub(r"[^a-z0-9]+", "_", text).strip("_")


def get_sheet(workbook, *candidates: str):
    for name in candidates:
        if name in workbook.sheetnames:
            return workbook[name]
    return None


def build_header_map(values: list[Any]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, value in enumerate(values):
        key = normalize_key(value)
        if key and key not in mapping:
            mapping[key] = index
    return mapping


def find_header_row(ws, required_keys: list[str], max_scan_rows: int = 10) -> tuple[int, dict[str, int]] | tuple[None, None]:
    for row_idx in range(1, min(ws.max_row, max_scan_rows) + 1):
        values = [ws.cell(row=row_idx, column=col).value for col in range(1, ws.max_column + 1)]
        header_map = build_header_map(values)
        if all(key in header_map for key in required_keys):
            return row_idx, header_map
    return None, None


def value_from_row(values: tuple[Any, ...], header_map: dict[str, int], *keys: str) -> Any:
    for key in keys:
        index = header_map.get(key)
        if index is not None and index < len(values):
            return values[index]
    return None


def month_number(value: Any) -> int | None:
    parsed = to_number(value)
    if isinstance(parsed, int) and 1 <= parsed <= 12:
        return parsed
    text = normalize_key(value)
    if not text:
        return None
    aliases = {
        "ene": 1,
        "enero": 1,
        "feb": 2,
        "febrero": 2,
        "mar": 3,
        "marzo": 3,
        "abr": 4,
        "abril": 4,
        "abri": 4,
        "may": 5,
        "mayo": 5,
        "jun": 6,
        "junio": 6,
        "jul": 7,
        "julio": 7,
        "ago": 8,
        "agosto": 8,
        "sep": 9,
        "sept": 9,
        "septiembre": 9,
        "oct": 10,
        "octubre": 10,
        "nov": 11,
        "noviembre": 11,
        "dic": 12,
        "diciembre": 12,
    }
    return aliases.get(text)


def trim_records(records: list[dict[str, Any]], limit: int) -> list[dict[str, Any]]:
    return records[:limit] if len(records) > limit else records


def load_client_rows(workbook) -> list[dict[str, Any]]:
    ws = get_sheet(workbook, "Datos clientes")
    if ws is None:
        return []

    header_row, header_map = find_header_row(ws, ["fecha_de_salida", "nombre", "producto"])
    if header_row is None or header_map is None:
        return []

    rows: list[dict[str, Any]] = []
    blank_streak = 0

    for values in ws.iter_rows(min_row=header_row + 1, max_col=ws.max_column, values_only=True):
        if not any(cell not in (None, "") for cell in values):
            if rows:
                blank_streak += 1
                if blank_streak >= 25:
                    break
            continue

        blank_streak = 0
        fecha_raw = value_from_row(values, header_map, "fecha_de_salida")
        fecha_iso = to_iso_date(fecha_raw)
        cantidad = int(to_number(value_from_row(values, header_map, "cantidad")) or 0)
        total = to_float(value_from_row(values, header_map, "total"))
        costo_unitario = to_float(value_from_row(values, header_map, "costo_producto"))

        row = {
            "fecha_salida": fecha_iso,
            "dia": to_number(value_from_row(values, header_map, "dia")),
            "mes": to_number(value_from_row(values, header_map, "mes")),
            "mes_letra": to_text(value_from_row(values, header_map, "mes_letra")),
            "anio": to_number(value_from_row(values, header_map, "ano")),
            "periodo": to_text(value_from_row(values, header_map, "periodo")),
            "nombre": to_text(value_from_row(values, header_map, "nombre")),
            "transportadora": to_text(value_from_row(values, header_map, "transportadora")),
            "numero_guia": identifier_text(value_from_row(values, header_map, "numero_de_guia", "numero_guia")),
            "direccion": to_text(value_from_row(values, header_map, "direccion")),
            "destino": to_text(value_from_row(values, header_map, "destino")),
            "canal_venta": to_text(value_from_row(values, header_map, "canal_de_venta")),
            "telefono": identifier_text(value_from_row(values, header_map, "telefono")),
            "cedula": identifier_text(value_from_row(values, header_map, "cedula")),
            "correo": to_text(value_from_row(values, header_map, "correo")),
            "producto": to_text(value_from_row(values, header_map, "producto")),
            "costo_producto": costo_unitario,
            "cantidad": cantidad,
            # La columna "Costo producto" en Datos clientes ya representa el costo total por fila;
            # sumar la columna tal cual es la fuente de verdad del usuario (no se multiplica por cantidad).
            "costo_total": costo_unitario,
            "total": total,
            "estado_envio": normalize_status(value_from_row(values, header_map, "estado_del_envio")),
        }

        if not row["fecha_salida"] and not row["nombre"] and not row["producto"] and not row["total"]:
            continue

        rows.append(row)

    return rows


def load_financial_rows(workbook) -> list[dict[str, Any]]:
    ws = get_sheet(workbook, "Total ingresado")
    if ws is None:
        return []

    header_row, header_map = find_header_row(
        ws,
        ["fecha", "utilidad_neta", "roas"],
    )
    if header_row is None or header_map is None:
        return []

    rows: list[dict[str, Any]] = []
    blank_streak = 0

    for values in ws.iter_rows(min_row=header_row + 1, max_col=ws.max_column, values_only=True):
        if not any(cell not in (None, "") for cell in values):
            if rows:
                blank_streak += 1
                if blank_streak >= 25:
                    break
            continue

        blank_streak = 0
        fecha_raw = value_from_row(values, header_map, "fecha")
        total_ingresado = to_float(
            value_from_row(
                values,
                header_map,
                "total_ingresado_logisto_99envios_y_picap",
                "ventas_brutas",
                "ventas_netas",
            )
        )
        utilidad_neta = to_float(value_from_row(values, header_map, "utilidad_neta"))

        row = {
            "fecha": to_iso_date(fecha_raw),
            "anio": to_number(value_from_row(values, header_map, "ano")),
            "total_ingresado": total_ingresado,
            "devoluciones": to_float(value_from_row(values, header_map, "devoluciones")),
            "gasto_publicidad": to_float(value_from_row(values, header_map, "gasto_en_publicidad", "gasto_publicidad")),
            "costo_producto": to_float(value_from_row(values, header_map, "costo_del_producto", "costo_producto")),
            "inversion_total": to_float(value_from_row(values, header_map, "inversion_total")),
            "utilidad_neta": utilidad_neta,
            "roas": round(to_float(value_from_row(values, header_map, "roas")), 2),
            "roi": round(to_float(value_from_row(values, header_map, "roi", "roi_pct")), 2),
            "margen_neto_pct": round(
                normalize_margin_pct(value_from_row(values, header_map, "margen_neto", "margen_neto_pct")),
                2,
            ),
        }

        if not row["fecha"] and not row["anio"] and not row["total_ingresado"] and not row["utilidad_neta"]:
            continue

        rows.append(row)

    rows.sort(key=lambda item: item["fecha"] or "")
    return rows


def load_retiro_rows(workbook) -> list[dict[str, Any]]:
    ws = get_sheet(workbook, "RENTABILIDAD POR RETIRO")
    rows: list[dict[str, Any]] = []
    if ws is not None:
        header_row, header_map = find_header_row(ws, ["fecha", "utilidad_neta", "roas"])
        if header_row is None or header_map is None:
            return []

        blank_streak = 0
        for values in ws.iter_rows(min_row=header_row + 1, max_col=ws.max_column, values_only=True):
            if not any(cell not in (None, "") for cell in values):
                if rows:
                    blank_streak += 1
                    if blank_streak >= 20:
                        break
                continue

            blank_streak = 0
            row = {
                "campania": to_text(value_from_row(values, header_map, "campana")),
                "fecha": to_iso_date(value_from_row(values, header_map, "fecha")),
                "mes": to_text(value_from_row(values, header_map, "mes")),
                "anio": to_number(value_from_row(values, header_map, "ano")),
                "total_ingresado": to_float(value_from_row(values, header_map, "total_ingresado")),
                "costo_producto": to_float(value_from_row(values, header_map, "costo_producto")),
                "devoluciones": to_float(value_from_row(values, header_map, "devoluciones")),
                "gasto_publicidad": to_float(value_from_row(values, header_map, "gasto_publicidad")),
                "utilidad_neta": to_float(value_from_row(values, header_map, "utilidad_neta")),
                "inversion_total": to_float(value_from_row(values, header_map, "inversion_total")),
                "posible_inversion_publicidad": to_float(
                    value_from_row(values, header_map, "posible_inversion_en_publicidad")
                ),
                "roas": round(to_float(value_from_row(values, header_map, "roas")), 2),
                "roi": round(to_float(value_from_row(values, header_map, "roi", "roi_pct")), 2),
                "margen_neto_pct": round(
                    normalize_margin_pct(value_from_row(values, header_map, "margen_neto", "margen_neto_pct")),
                    2,
                ),
            }

            if not row["campania"] and not row["fecha"] and not row["total_ingresado"]:
                continue

            rows.append(row)

        rows.sort(key=lambda item: item["fecha"] or "")
        return rows

    ws = get_sheet(workbook, "Análisis Ventas & Publicidad")
    if ws is None:
        return []

    header_row, header_map = find_header_row(ws, ["ano", "mes", "utilidad_neta", "roas"], max_scan_rows=8)
    if header_row is None or header_map is None:
        return []

    blank_streak = 0
    for values in ws.iter_rows(min_row=header_row + 1, max_col=ws.max_column, values_only=True):
        if not any(cell not in (None, "") for cell in values):
            if rows:
                blank_streak += 1
                if blank_streak >= 20:
                    break
            continue

        blank_streak = 0
        anio = to_number(value_from_row(values, header_map, "ano"))
        mes_raw = value_from_row(values, header_map, "mes")
        mes_num = month_number(mes_raw)
        fecha_iso = None
        if isinstance(anio, int) and mes_num:
            fecha_iso = date(anio, mes_num, 1).isoformat()

        gasto_publicidad = to_float(value_from_row(values, header_map, "gasto_publicidad"))
        costo_producto = to_float(value_from_row(values, header_map, "costo_producto"))
        row = {
            "campania": "General",
            "fecha": fecha_iso,
            "mes": to_text(mes_raw),
            "anio": anio,
            "total_ingresado": to_float(
                value_from_row(values, header_map, "ventas_netas", "ventas_brutas")
            ),
            "costo_producto": costo_producto,
            "devoluciones": to_float(value_from_row(values, header_map, "devoluciones")),
            "gasto_publicidad": gasto_publicidad,
            "utilidad_neta": to_float(value_from_row(values, header_map, "utilidad_neta")),
            "inversion_total": gasto_publicidad + costo_producto,
            "posible_inversion_publicidad": 0.0,
            "roas": round(to_float(value_from_row(values, header_map, "roas")), 2),
            "roi": round(to_float(value_from_row(values, header_map, "roi", "roi_pct")), 2),
            "margen_neto_pct": round(
                normalize_margin_pct(value_from_row(values, header_map, "margen_neto", "margen_neto_pct")),
                2,
            ),
        }

        if not row["fecha"] and not row["total_ingresado"] and not row["utilidad_neta"]:
            continue

        rows.append(row)

    rows.sort(key=lambda item: item["fecha"] or "")
    return rows


def load_publicidad_rows(workbook) -> list[dict[str, Any]]:
    ws = get_sheet(workbook, "Gasto en publicidad", "Data publicidad")
    if ws is None:
        return []

    header_row, header_map = find_header_row(ws, ["fecha", "pago"])
    if header_row is None or header_map is None:
        return []

    rows: list[dict[str, Any]] = []
    blank_streak = 0

    for values in ws.iter_rows(min_row=header_row + 1, max_col=ws.max_column, values_only=True):
        if not any(cell not in (None, "") for cell in values):
            if rows:
                blank_streak += 1
                if blank_streak >= 25:
                    break
            continue

        blank_streak = 0
        row = {
            "fecha": to_iso_date(value_from_row(values, header_map, "fecha", "fechacompleta")),
            "anio": to_number(value_from_row(values, header_map, "ano")),
            "mes": to_text(value_from_row(values, header_map, "mes")),
            "pago": to_float(value_from_row(values, header_map, "pago")),
        }

        if not row["fecha"] and not row["anio"] and not row["pago"]:
            continue

        rows.append(row)

    rows.sort(key=lambda item: item["fecha"] or "")
    return rows


def load_inventory_rows(workbook) -> dict[str, list[dict[str, Any]]]:
    ws = get_sheet(workbook, "Inventario sets y tiaras solas")
    sets_rows: list[dict[str, Any]] = []
    tiaras_rows: list[dict[str, Any]] = []
    if ws is not None:
        blank_streak = 0
        for values in ws.iter_rows(min_row=3, max_col=15, values_only=True):
            set_name = to_text(values[0])
            tiara_name = to_text(values[8])

            if not set_name and not tiara_name:
                if sets_rows or tiaras_rows:
                    blank_streak += 1
                    if blank_streak >= 10:
                        break
                continue

            blank_streak = 0

            if set_name:
                cantidad = int(to_number(values[1]) or 0)
                costo_producto = to_float(values[2])
                precio_venta = to_float(values[5])
                sets_rows.append(
                    {
                        "producto": set_name,
                        "cantidad": cantidad,
                        "costo_producto": costo_producto,
                        "salidas": to_number(values[3]) or 0,
                        "codigo": to_text(values[4]),
                        "precio_venta": precio_venta,
                        "valor_unidades_vendidas": to_float(values[6]) or precio_venta * cantidad,
                        "valor_potencial": precio_venta * cantidad,
                    }
                )

            if tiara_name:
                cantidad = int(to_number(values[9]) or 0)
                costo_producto = to_float(values[10])
                precio_venta = to_float(values[13])
                tiaras_rows.append(
                    {
                        "producto": tiara_name,
                        "cantidad": cantidad,
                        "costo_producto": costo_producto,
                        "salidas": to_number(values[11]) or 0,
                        "codigo": to_text(values[12]),
                        "precio_venta": precio_venta,
                        "valor_unidades_vendidas": to_float(values[14]) or precio_venta * cantidad,
                        "valor_potencial": precio_venta * cantidad,
                    }
                )

        return {
            "sets": sets_rows,
            "tiaras": tiaras_rows,
        }

    ws = get_sheet(workbook, "Catálogo productos")
    if ws is None:
        return {
            "sets": [],
            "tiaras": [],
        }

    for values in ws.iter_rows(min_row=5, max_col=6, values_only=True):
        producto = to_text(values[0])
        if not producto:
            continue
        costo_unitario = to_float(values[1])
        cantidad = int(to_number(values[2]) or 0)
        item = {
            "producto": producto,
            "cantidad": cantidad,
            "costo_producto": costo_unitario,
            "salidas": cantidad,
            "codigo": None,
            "precio_venta": 0.0,
            "valor_unidades_vendidas": costo_unitario * cantidad,
            "valor_potencial": costo_unitario * cantidad,
        }
        normalized_name = normalize_key(producto)
        if "tiara" in normalized_name or "corona" in normalized_name:
            tiaras_rows.append(item)
        else:
            sets_rows.append(item)

    return {
        "sets": sets_rows,
        "tiaras": tiaras_rows,
    }


def group_clientes_by_month(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for item in records:
        fecha_iso = item.get("fecha_salida")
        if not fecha_iso:
            continue
        dt = date.fromisoformat(fecha_iso)
        key = f"{dt.year}-{dt.month:02d}"
        bucket = grouped.setdefault(
            key,
            {
                "periodo": key,
                "label": month_label(dt),
                "total_envios": 0,
                "valor_total": 0.0,
                "entregados": 0,
            },
        )
        bucket["total_envios"] += 1
        bucket["valor_total"] += item.get("total", 0.0) or 0.0
        estado = (item.get("estado_envio") or "").lower()
        if "entregado" in estado:
            bucket["entregados"] += 1

    ordered = [grouped[key] for key in sorted(grouped)]
    for item in ordered:
        item["pendientes"] = item["total_envios"] - item["entregados"]
        item["valor_total"] = round(item["valor_total"], 2)
    return ordered


def group_transportadoras(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for item in records:
        name = item.get("transportadora") or "Sin transportadora"
        bucket = grouped.setdefault(name, {"transportadora": name, "cantidad": 0, "valor_total": 0.0})
        bucket["cantidad"] += 1
        bucket["valor_total"] += item.get("total", 0.0) or 0.0
    return sorted(grouped.values(), key=lambda entry: (-entry["cantidad"], -entry["valor_total"]))[:8]


def group_statuses(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    counter = Counter(item.get("estado_envio") or "Sin estado" for item in records)
    return [{"estado": key, "cantidad": value} for key, value in counter.most_common()]


def group_productos_por_periodo(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[int, int, str], dict[str, Any]] = {}
    for item in records:
        producto = item.get("producto")
        if not producto:
            continue

        anio = item.get("anio")
        mes = item.get("mes")
        if anio is None or mes is None:
            fecha_iso = item.get("fecha_salida")
            if fecha_iso:
                try:
                    dt = date.fromisoformat(fecha_iso)
                    anio = dt.year
                    mes = dt.month
                except ValueError:
                    pass

        try:
            anio_int = int(anio) if anio is not None else None
            mes_int = int(mes) if mes is not None else None
        except (TypeError, ValueError):
            continue

        if anio_int is None or mes_int is None or not 1 <= mes_int <= 12:
            continue

        key = (anio_int, mes_int, producto)
        bucket = grouped.setdefault(
            key,
            {
                "anio": anio_int,
                "num_mes": mes_int,
                "producto": producto,
                "cantidad": 0,
                "total": 0.0,
                "costo_total": 0.0,
                "ordenes": 0,
            },
        )
        bucket["cantidad"] += int(item.get("cantidad") or 0)
        bucket["total"] += item.get("total", 0.0) or 0.0
        bucket["costo_total"] += item.get("costo_total", 0.0) or 0.0
        bucket["ordenes"] += 1

    result = list(grouped.values())
    result.sort(key=lambda r: (r["anio"], r["num_mes"], -r["cantidad"]))
    return result


def group_publicidad_by_month(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for item in records:
        fecha_iso = item.get("fecha")
        if not fecha_iso:
            continue
        dt = date.fromisoformat(fecha_iso)
        key = f"{dt.year}-{dt.month:02d}"
        bucket = grouped.setdefault(
            key,
            {
                "periodo": key,
                "label": month_label(dt),
                "pagos": 0,
                "total_pago": 0.0,
            },
        )
        bucket["pagos"] += 1
        bucket["total_pago"] += item.get("pago", 0.0) or 0.0
    ordered = [grouped[key] for key in sorted(grouped)]
    for item in ordered:
        item["total_pago"] = round(item["total_pago"], 2)
    return ordered


def summarize_inventory(items: list[dict[str, Any]]) -> dict[str, Any]:
    unidades = sum(int(item.get("cantidad") or 0) for item in items)
    costo_total = sum((item.get("costo_producto") or 0.0) * (item.get("cantidad") or 0) for item in items)
    valor_potencial = sum(item.get("valor_potencial") or 0.0 for item in items)
    return {
        "referencias": len(items),
        "unidades": unidades,
        "costo_total": round(costo_total, 2),
        "valor_potencial": round(valor_potencial, 2),
        "top_valor": sorted(items, key=lambda item: item.get("valor_potencial", 0.0), reverse=True)[:5],
    }


def summarize(workbook_path: Path) -> dict[str, Any]:
    with open_workbook_safely(workbook_path) as workbook:
        clientes = load_client_rows(workbook)
        finanzas = load_financial_rows(workbook)
        retiro = load_retiro_rows(workbook)
        publicidad = load_publicidad_rows(workbook)
        inventario = load_inventory_rows(workbook)
        sheet_names = workbook.sheetnames

    valor_total_clientes = round(sum(item.get("total", 0.0) or 0.0 for item in clientes), 2)
    costo_total_clientes = round(sum(item.get("costo_total", 0.0) or 0.0 for item in clientes), 2)
    clientes_unicos = len(
        {
            (item.get("nombre") or "").strip().casefold()
            for item in clientes
            if item.get("nombre")
        }
    )
    entregados = sum(
        1 for item in clientes if "entregado" in (item.get("estado_envio") or "").lower()
    )

    total_ingresado = round(sum(item.get("total_ingresado", 0.0) for item in finanzas), 2)
    total_utilidad = round(sum(item.get("utilidad_neta", 0.0) for item in finanzas), 2)
    total_inversion = round(sum(item.get("inversion_total", 0.0) for item in finanzas), 2)
    total_publicidad = round(sum(item.get("gasto_publicidad", 0.0) for item in finanzas), 2)

    retiro_total_ingresado = round(sum(item.get("total_ingresado", 0.0) for item in retiro), 2)
    retiro_total_utilidad = round(sum(item.get("utilidad_neta", 0.0) for item in retiro), 2)
    retiro_total_publicidad = round(sum(item.get("gasto_publicidad", 0.0) for item in retiro), 2)
    mejor_campania = max(retiro, key=lambda item: item.get("utilidad_neta", 0.0), default=None)

    inventario_sets = summarize_inventory(inventario["sets"])
    inventario_tiaras = summarize_inventory(inventario["tiaras"])

    clientes_recientes = sorted(
        clientes,
        key=lambda item: item.get("fecha_salida") or "",
        reverse=True,
    )[:20]

    output = {
        "metadata": {
            "sourcePath": str(workbook_path),
            "fileName": workbook_path.name,
            "lastModified": datetime.fromtimestamp(workbook_path.stat().st_mtime).isoformat(),
            "sheetNames": sheet_names,
            "extractedAt": datetime.utcnow().isoformat() + "Z",
        },
        "overview": {
            "clientes": {
                "registros": len(clientes),
                "clientes_unicos": clientes_unicos,
                "valor_total": valor_total_clientes,
                "costo_total": costo_total_clientes,
                "ticket_promedio": round(valor_total_clientes / len(clientes), 2) if clientes else 0,
                "entregados": entregados,
                "pendientes": max(len(clientes) - entregados, 0),
                "transportadoras_activas": len(
                    {item.get("transportadora") for item in clientes if item.get("transportadora")}
                ),
            },
            "contabilidad": {
                "periodos": len(finanzas),
                "total_ingresado": total_ingresado,
                "utilidad_neta": total_utilidad,
                "inversion_total": total_inversion,
                "gasto_publicidad": total_publicidad,
                "devoluciones": round(sum(item.get("devoluciones", 0.0) for item in finanzas), 2),
                "roas_promedio": round(
                    sum(item.get("roas", 0.0) for item in finanzas) / len(finanzas), 2
                )
                if finanzas
                else 0,
                "roi_promedio": round(
                    sum(item.get("roi", 0.0) for item in finanzas) / len(finanzas), 2
                )
                if finanzas
                else 0,
                "margen_neto_promedio": round(
                    sum(item.get("margen_neto_pct", 0.0) for item in finanzas) / len(finanzas), 2
                )
                if finanzas
                else 0,
            },
            "retiro": {
                "registros": len(retiro),
                "total_ingresado": retiro_total_ingresado,
                "utilidad_neta": retiro_total_utilidad,
                "gasto_publicidad": retiro_total_publicidad,
                "margen_neto_promedio": round(
                    sum(item.get("margen_neto_pct", 0.0) for item in retiro) / len(retiro), 2
                )
                if retiro
                else 0,
                "mejor_campania": mejor_campania,
            },
            "inventario": {
                "total_referencias": inventario_sets["referencias"] + inventario_tiaras["referencias"],
                "total_unidades": inventario_sets["unidades"] + inventario_tiaras["unidades"],
                "costo_total": round(
                    inventario_sets["costo_total"] + inventario_tiaras["costo_total"], 2
                ),
                "valor_potencial": round(
                    inventario_sets["valor_potencial"] + inventario_tiaras["valor_potencial"], 2
                ),
            },
        },
        "enviosPorMes": group_clientes_by_month(clientes),
        "estadosEnvio": group_statuses(clientes),
        "transportadoras": group_transportadoras(clientes),
        "productosPorPeriodo": group_productos_por_periodo(clientes),
        "clientesRecientes": clientes_recientes,
        "finanzasPorPeriodo": finanzas,
        "rentabilidadRetiro": retiro,
        "publicidadPorMes": group_publicidad_by_month(publicidad),
        "inventario": {
            "sets": inventario["sets"],
            "tiaras": inventario["tiaras"],
            "resumen": {
                "sets": inventario_sets,
                "tiaras": inventario_tiaras,
            },
        },
    }

    return json.loads(json.dumps(output, default=to_serialized))


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--path", default=str(DEFAULT_EXCEL_PATH))
    args = parser.parse_args()

    workbook_path = Path(args.path).expanduser()
    if not workbook_path.exists():
        raise FileNotFoundError(f"No se encontró el archivo Excel: {workbook_path}")

    payload = summarize(workbook_path)
    print(json.dumps(payload, ensure_ascii=False))


if __name__ == "__main__":
    main()
