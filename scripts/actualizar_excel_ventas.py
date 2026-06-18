"""
Descarga ordenes de MeLi desde la ultima fecha del Excel hasta hoy,
las formatea igual que el historico y las agrega al Excel.

Uso:
    python scripts/actualizar_excel_ventas.py
    python scripts/actualizar_excel_ventas.py --desde 2026-06-01 --hasta 2026-06-05
"""
import argparse
import sys
import json
import urllib.request
import urllib.parse
import urllib.error
from datetime import datetime, timezone, timedelta
from pathlib import Path

import os
import psycopg2
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv

# Cargar variables desde el .env del backend
load_dotenv(dotenv_path=Path(__file__).resolve().parents[1] / "sepia meli api" / ".env")

# ── Configuracion ──────────────────────────────────────────────────────────
EXCEL_PATH = Path(r"C:\Users\SANTIAGO\One Drive\OneDrive\Excel sepia\Mercado LibreOFICIAL .xlsx")
SHEET_NAME = "Data"

DB_HOST = os.getenv("DB_HOST", "127.0.0.1")
DB_PORT = int(os.getenv("DB_PORT", "5432"))
DB_USER = os.getenv("DB_USER", "postgres")
DB_PASSWORD = os.getenv("DB_PASSWORD", "")
DB_NAME = os.getenv("DB_NAME", "mercado_libre_oficial")

MELI_CLIENT_ID = os.getenv("MELI_CLIENT_ID", "")
MELI_CLIENT_SECRET = os.getenv("MELI_CLIENT_SECRET", "")
MELI_TOKEN_URL = "https://api.mercadolibre.com/oauth/token"
MELI_API = "https://api.mercadolibre.com"

MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}
ESTADOS_ES = {
    "paid": "Entregado",
    "delivered": "Entregado",
    "confirmed": "Confirmado",
    "cancelled": "Cancelado",
    "pending": "Pendiente",
    "invalid": "Invalido",
}

# ── Helpers HTTP ───────────────────────────────────────────────────────────

def _http_post(url, data):
    body = urllib.parse.urlencode(data).encode()
    req = urllib.request.Request(url, data=body, method="POST")
    req.add_header("Content-Type", "application/x-www-form-urlencoded")
    req.add_header("Accept", "application/json")
    with urllib.request.urlopen(req, timeout=30) as resp:
        return json.loads(resp.read())


def _http_get(url, token, params=None):
    if params:
        url = url + "?" + urllib.parse.urlencode(params)
    req = urllib.request.Request(url)
    req.add_header("Authorization", f"Bearer {token}")
    req.add_header("Accept", "application/json")
    with urllib.request.urlopen(req, timeout=30) as resp:
        return json.loads(resp.read())


# ── Tokens ─────────────────────────────────────────────────────────────────

def get_db_conn():
    return psycopg2.connect(
        host=DB_HOST, port=DB_PORT, user=DB_USER,
        password=DB_PASSWORD, dbname=DB_NAME,
    )


def load_token():
    conn = get_db_conn()
    cur = conn.cursor()
    cur.execute("SELECT access_token, refresh_token, expires_at FROM meli_tokens ORDER BY id LIMIT 1")
    row = cur.fetchone()
    conn.close()
    if not row:
        sys.exit("ERROR: No hay tokens en la BD. Inicia sesion en el dashboard primero.")
    return {"access_token": row[0], "refresh_token": row[1], "expires_at": row[2]}


def refresh_token(refresh_tok):
    print("  Refrescando token de acceso...")
    data = {
        "grant_type": "refresh_token",
        "client_id": MELI_CLIENT_ID,
        "client_secret": MELI_CLIENT_SECRET,
        "refresh_token": refresh_tok,
    }
    resp = _http_post(MELI_TOKEN_URL, data)
    new_access = resp["access_token"]
    new_refresh = resp.get("refresh_token", refresh_tok)
    expires_in = resp.get("expires_in", 21600)
    expires_at = datetime.now(timezone.utc) + timedelta(seconds=expires_in)

    conn = get_db_conn()
    cur = conn.cursor()
    cur.execute(
        "UPDATE meli_tokens SET access_token=%s, refresh_token=%s, expires_at=%s, updated_at=NOW()",
        (new_access, new_refresh, expires_at),
    )
    conn.commit()
    conn.close()
    print(f"  Token renovado. Expira: {expires_at.strftime('%Y-%m-%d %H:%M UTC')}")
    return new_access


def get_valid_token():
    tok = load_token()
    now = datetime.now(timezone.utc)
    exp = tok["expires_at"]
    if exp.tzinfo is None:
        exp = exp.replace(tzinfo=timezone.utc)
    if exp - now < timedelta(minutes=5):
        return refresh_token(tok["refresh_token"])
    return tok["access_token"]


# ── API MeLi ───────────────────────────────────────────────────────────────

def get_seller_id(token):
    data = _http_get(f"{MELI_API}/users/me", token)
    return str(data["id"])


def get_orders_range(token, seller_id, date_from, date_to):
    """Devuelve lista de ordenes entre date_from y date_to (strings YYYY-MM-DD)."""
    orders = []
    offset = 0
    limit = 50
    date_from_iso = f"{date_from}T00:00:00.000-05:00"
    date_to_iso = f"{date_to}T23:59:59.000-05:00"

    print(f"  Descargando ordenes {date_from} → {date_to}...")
    while True:
        params = {
            "seller": seller_id,
            "order.date_created.from": date_from_iso,
            "order.date_created.to": date_to_iso,
            "offset": offset,
            "limit": limit,
            "sort": "date_asc",
        }
        try:
            resp = _http_get(f"{MELI_API}/orders/search", token, params)
        except urllib.error.HTTPError as e:
            print(f"  ERROR HTTP {e.code}: {e.read()[:300]}")
            break

        results = resp.get("results", [])
        orders.extend(results)

        paging = resp.get("paging", {})
        total = paging.get("total", 0)
        offset += len(results)
        print(f"  ... {len(orders)}/{total} ordenes")
        if not results or offset >= total:
            break

    return orders


def get_item_details(token, item_id, cache={}):
    if item_id in cache:
        return cache[item_id]
    try:
        data = _http_get(f"{MELI_API}/items/{item_id}", token)
        cache[item_id] = data
        return data
    except Exception:
        return {}


def get_category_name(token, cat_id, cache={}):
    if cat_id in cache:
        return cache[cat_id]
    try:
        data = _http_get(f"{MELI_API}/categories/{cat_id}", token)
        name = data.get("name", cat_id)
        cache[cat_id] = name
        return name
    except Exception:
        return cat_id


# ── Mapeo de orden MeLi → fila Excel ──────────────────────────────────────

def order_to_rows(order, token):
    """Una orden puede tener multiples items -> multiples filas."""
    rows = []

    date_str = order.get("date_closed") or order.get("date_created", "")
    try:
        dt = datetime.fromisoformat(date_str.replace("Z", "+00:00"))
        # Convertir a hora Colombia (UTC-5)
        dt_col = dt.astimezone(timezone(timedelta(hours=-5)))
    except Exception:
        dt_col = datetime.now()

    fecha = dt_col.strftime("%d/%m/%Y")
    anio = dt_col.year
    mes_num = dt_col.month
    mes_nom = MESES_ES[mes_num]
    dia = dt_col.day

    numero_venta = str(order.get("id", ""))
    estado_raw = order.get("status", "")
    estado = ESTADOS_ES.get(estado_raw, estado_raw.capitalize())

    buyer = order.get("buyer", {})
    comprador = f"{buyer.get('first_name', '')} {buyer.get('last_name', '')}".strip()

    shipping = order.get("shipping", {})
    shipping_id = shipping.get("id")

    # Intentar obtener ciudad del shipping
    ciudad = ""
    if shipping_id:
        try:
            ship_data = _http_get(f"{MELI_API}/shipments/{shipping_id}", token)
            dest = ship_data.get("receiver_address", {})
            ciudad = dest.get("city", {}).get("name", "") if isinstance(dest.get("city"), dict) else ""
            logistic = ship_data.get("logistic_type", "")
            if "drop_off" in logistic or "self_service" in logistic:
                forma_entrega = "Correo y puntos de despacho"
            elif "xd_drop_off" in logistic:
                forma_entrega = "Punto Meli"
            elif "home_delivery" in logistic or "fulfillment" in logistic:
                forma_entrega = "Domicilio"
            else:
                forma_entrega = logistic or "Correo y puntos de despacho"
        except Exception:
            forma_entrega = "Correo y puntos de despacho"
    else:
        forma_entrega = "Correo y puntos de despacho"

    payments = order.get("payments", [])
    monto_reportado = sum(float(p.get("total_paid_amount", 0)) for p in payments if p.get("status") == "approved")
    cargo_venta_imp = sum(float(p.get("marketplace_fee", 0) or 0) for p in payments)

    # Recorrer items de la orden
    order_items = order.get("order_items", [])
    for oi in order_items:
        item = oi.get("item", {})
        item_id = item.get("id", "")
        titulo = item.get("title", "")
        cantidad = int(oi.get("quantity", 1))
        precio_unitario = float(oi.get("unit_price", 0))

        # Variante (atributos de la variacion)
        variation_id = oi.get("variation_id")
        variante_talla = ""
        genero = ""
        categoria = ""

        item_det = {}
        if item_id:
            item_det = get_item_details(token, item_id)
            cat_id = item_det.get("category_id", "")
            if cat_id:
                categoria = get_category_name(token, cat_id)

            # Buscar variante
            if variation_id and item_det.get("variations"):
                for var in item_det["variations"]:
                    if str(var.get("id")) == str(variation_id):
                        attrs = {a["id"]: a.get("value_name", "") for a in var.get("attribute_combinations", [])}
                        talla = attrs.get("SIZE", attrs.get("CLOTHING_SIZE", attrs.get("SHOE_SIZE", "")))
                        col = attrs.get("COLOR", "")
                        acabado = attrs.get("ACABADO", "")
                        parts = []
                        if acabado:
                            parts.append(f"Acabado : {acabado}")
                        if col:
                            parts.append(f"Color : {col}")
                        if talla:
                            parts.append(f"Talla : {talla}")
                        variante_talla = " | ".join(parts) if parts else ""

                        # Genero desde atributos del item principal
                        for a in item_det.get("attributes", []):
                            if a.get("id") == "GENDER":
                                g = a.get("value_name", "")
                                if "femen" in g.lower() or "dama" in g.lower() or "mujer" in g.lower():
                                    genero = "Dama"
                                elif "mascul" in g.lower() or "hombre" in g.lower() or "caball" in g.lower():
                                    genero = "Caballero"
                                else:
                                    genero = g
                        break

        # Ingresos de envio y costo de envio: distribuir proporcionalmente
        n = len(order_items) or 1
        ingresos_envio = 0.0
        costos_envio = round(-abs(float(order.get("shipping_amount", 0) or 0)) / n, 2)

        row = {
            "Año": anio,
            "Mes": mes_nom,
            "Num_mes": mes_num,
            "Día": dia,
            "Fecha": fecha,
            "Numero_venta": numero_venta,
            "Estado": estado,
            "Producto": titulo,
            "Categoria": categoria,
            "Variante_Talla": variante_talla,
            "Genero": genero,
            "Cantidad": cantidad,
            "Monto_reportado_COP": round(monto_reportado / n, 2) if n > 1 else monto_reportado,
            "Ingresos_productos_COP": round(precio_unitario * cantidad, 2),
            "Cargo_venta_impuestos_COP": round(-abs(cargo_venta_imp) / n, 2),
            "Ingresos_envio_COP": ingresos_envio if ingresos_envio else None,
            "Costos_envio_COP": costos_envio if costos_envio else None,
            "Anulaciones_reembolsos_COP": None,
            "Publicacion_ID": item_id,
            "Precio_unitario_publicacion_COP": precio_unitario,
            "Comprador": comprador,
            "Ciudad": ciudad,
            "Forma_entrega": forma_entrega,
        }
        rows.append(row)

    return rows


# ── Excel ──────────────────────────────────────────────────────────────────

COLUMNAS = [
    "Año", "Mes", "Num_mes", "Día", "Fecha", "Numero_venta", "Estado",
    "Producto", "Categoria", "Variante_Talla", "Genero", "Cantidad",
    "Monto_reportado_COP", "Ingresos_productos_COP", "Cargo_venta_impuestos_COP",
    "Ingresos_envio_COP", "Costos_envio_COP", "Anulaciones_reembolsos_COP",
    "Publicacion_ID", "Precio_unitario_publicacion_COP", "Comprador", "Ciudad",
    "Forma_entrega",
]

# Mapeo nombre columna Excel (con tildes) → nombre en nuestro dict
COL_MAP = {
    "Año": "Año", "A�o": "Año",
    "Mes": "Mes",
    "Num_mes": "Num_mes",
    "Día": "Día", "D�a": "Día",
    "Fecha": "Fecha",
    "Numero_venta": "Numero_venta",
    "Estado": "Estado",
    "Producto": "Producto",
    "Categoria": "Categoria",
    "Variante_Talla": "Variante_Talla",
    "Genero": "Genero",
    "Cantidad": "Cantidad",
    "Monto_reportado_COP": "Monto_reportado_COP",
    "Ingresos_productos_COP": "Ingresos_productos_COP",
    "Cargo_venta_impuestos_COP": "Cargo_venta_impuestos_COP",
    "Ingresos_envio_COP": "Ingresos_envio_COP",
    "Costos_envio_COP": "Costos_envio_COP",
    "Anulaciones_reembolsos_COP": "Anulaciones_reembolsos_COP",
    "Publicacion_ID": "Publicacion_ID",
    "Precio_unitario_publicacion_COP": "Precio_unitario_publicacion_COP",
    "Comprador": "Comprador",
    "Ciudad": "Ciudad",
    "Forma_entrega": "Forma_entrega",
}


def append_to_excel(new_rows, excel_path, sheet_name):
    print(f"\nAbriendo Excel: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)

    if sheet_name not in wb.sheetnames:
        print(f"  Hoja '{sheet_name}' no encontrada. Hojas disponibles: {wb.sheetnames}")
        sheet_name = wb.sheetnames[0]
        print(f"  Usando hoja: {sheet_name}")

    ws = wb[sheet_name]

    # Leer encabezados de la fila 1 (con posible encoding roto)
    headers_raw = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    # Mapear a nombres normalizados
    col_indices = {}  # nombre_normalizado -> indice 1-based
    for i, h in enumerate(headers_raw, start=1):
        if h is None:
            continue
        norm = COL_MAP.get(str(h), str(h))
        col_indices[norm] = i

    print(f"  Ultima fila con datos: {ws.max_row}")
    print(f"  Columnas detectadas: {list(col_indices.keys())[:6]}...")

    # Copiar formato de la ultima fila de datos para referencia
    last_row = ws.max_row

    rows_added = 0
    for row_data in new_rows:
        next_row = ws.max_row + 1
        for col_name, col_idx in col_indices.items():
            val = row_data.get(col_name)
            cell = ws.cell(row=next_row, column=col_idx, value=val)
            # Copiar numero de formato de la fila anterior
            src = ws.cell(row=last_row, column=col_idx)
            if src.number_format and src.number_format != "General":
                cell.number_format = src.number_format
        rows_added += 1

    print(f"  Agregadas {rows_added} filas nuevas.")
    print(f"  Guardando...")
    wb.save(excel_path)
    print(f"  Guardado OK.")
    return rows_added


# ── Main ───────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--desde", help="Fecha inicio YYYY-MM-DD (por defecto: dia siguiente al ultimo en Excel)")
    parser.add_argument("--hasta", help="Fecha fin YYYY-MM-DD (por defecto: hoy)")
    args = parser.parse_args()

    # Determinar rango de fechas
    if args.hasta:
        hasta = args.hasta
    else:
        hasta = datetime.now().strftime("%Y-%m-%d")

    if args.desde:
        desde = args.desde
    else:
        # Leer ultima fecha del Excel
        print("Leyendo Excel para determinar ultima fecha...")
        df = pd.read_excel(EXCEL_PATH, sheet_name=SHEET_NAME, usecols=["Fecha"])
        ultima = df["Fecha"].dropna().iloc[-1]
        if hasattr(ultima, "strftime"):
            ultima_dt = ultima
        else:
            ultima_dt = datetime.strptime(str(ultima), "%d/%m/%Y")
        siguiente = ultima_dt + timedelta(days=1)
        desde = siguiente.strftime("%Y-%m-%d")
        print(f"  Ultima fecha en Excel: {ultima_dt.strftime('%d/%m/%Y')}")

    print(f"\nRango a descargar: {desde} → {hasta}")

    if desde > hasta:
        print("El Excel ya esta al dia. No hay fechas nuevas que agregar.")
        return

    # Obtener token valido
    print("\nObteniendo token de acceso MeLi...")
    token = get_valid_token()
    print("  Token OK.")

    # Obtener seller ID
    print("Obteniendo ID del vendedor...")
    seller_id = get_seller_id(token)
    print(f"  Seller ID: {seller_id}")

    # Descargar ordenes
    orders = get_orders_range(token, seller_id, desde, hasta)
    print(f"\nTotal ordenes descargadas: {len(orders)}")

    if not orders:
        print("No hay ordenes nuevas en ese rango.")
        return

    # Convertir ordenes a filas
    print("\nProcesando ordenes (obteniendo detalles de items y envios)...")
    all_rows = []
    for i, order in enumerate(orders, 1):
        print(f"  [{i}/{len(orders)}] Orden {order.get('id')}")
        rows = order_to_rows(order, token)
        all_rows.extend(rows)

    print(f"\nTotal filas a agregar: {len(all_rows)}")

    # Agregar al Excel
    n = append_to_excel(all_rows, EXCEL_PATH, SHEET_NAME)
    print(f"\nListo. Se agregaron {n} filas al Excel.")
    print(f"Archivo: {EXCEL_PATH}")


if __name__ == "__main__":
    main()
